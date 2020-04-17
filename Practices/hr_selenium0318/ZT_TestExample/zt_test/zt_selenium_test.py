"""
@Author  : Laura
@File    : zt_login.py
@Time    : 2020/3/18 14:25
"""
'''
使用unittest+selenium 搭建ui自动化测试框架
1. 可以分模块写测试用例集：如首页模块单独作为一个测试集，写成一个测试类TestShouYe，继承测试主类（TestZt），主类继承unittest.TestCase
    其中主类TestZt主要用于优化重复可使用的函数、操作或参数的定义与存储，做通用类参数化
    测试类如TestShouYe根据测试网页定义，其中包含该模块下各项测试用例，一个函数定义一类测试用例，再使用参数化脚本结合openpyxl读取excel中测试数据，生成多条实例用例。
    setup()函数与teardown()函数，为unittest 框架特有函数，执行每条用例前执行，可用于
    类中初始化函数 __init__使用时注意超继承，避免被重写
2. 执行用例：使用unittest.TestSuit()创建存储器，使用 unittest.TestLoader()创建加载器加载用例，添加至存储器，使用HTMLTestRunner运行测试用例生成html测试报告
3. 参数化，方法一：使用初始化函数超继承，写入测试数据
            方法二：使用unittest + ddt ，测试用例函数即可以参数行驶传入测试数据   
'''


from selenium import webdriver
import unittest
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from HTMLTestRunner import HTMLTestRunner
import openpyxl


class TestZt(unittest.TestCase):

    # @classmethod   #类方法修饰器
    # def setUpClass(cls):  #必须使用@classmethod 装饰器,  所有case运行之前只运行一次    setUp():每个测试case运行之前运行
    #     #配置驱动path
    #     # chrome_driver = r"E:\Program Files\python 3.8.2\Lib\site-packages\selenium\chromedriver.exe"
    #     # 配置浏览器 浏览器中输入chrome://version/ 查看个人资料路径配置
    #     # option = webdriver.ChromeOptions()
    #     # option.add_argument(r'--user-data-dir = \Users\admin\AppData\Local\Google\Chrome\User Data\Profile 1')
    #     # self.driver = webdriver.Chrome(executable_path=chrome_driver, chrome_options=option)
    #     cls.driver = webdriver.Chrome()
    #     cls.driver.implicitly_wait(5)
    #     cls.url = "http://zt.emhes.cn/"
    #     # self.verificationErrors = []
    #     global driver
    #     driver = cls.driver
    #     #进入首页并登陆
    #     driver.get(cls.url)
    #     time.sleep(1)
    #     driver.find_element_by_id("userName").send_keys("hwy")
    #     driver.find_element_by_id("password").send_keys("123456")
    #     driver.find_element_by_id("loginBtn").click()
    #     # 浏览器窗口最大化
    #     driver.maximize_window()
    #
    # @classmethod
    # def tearDownclass(cls): #必须使用@classmethod装饰器, 所有case运行完之后只运行一次   tearDown():每个测试case运行完之后执行
    #     #关闭浏览器
    #     driver.quit()




    def setUp(self):  #必须使用@classmethod 装饰器,  所有case运行之前只运行一次    setUp():每个测试case运行之前运行。。。初始化函数
        #配置驱动path
        # chrome_driver = r"E:\Program Files\python 3.8.2\Lib\site-packages\selenium\chromedriver.exe"
        # 配置浏览器 浏览器中输入chrome://version/ 查看个人资料路径配置
        option = webdriver.ChromeOptions()
        option.add_argument(r'--user-data-dir = \Users\admin\AppData\Local\Google\Chrome\User Data\Profile 1')
        # self.driver = webdriver.Chrome(executable_path=chrome_driver, chrome_options=option)
        self.driver = webdriver.Chrome(chrome_options=option)
        self.driver.implicitly_wait(5)
        self.url = "http://zt.emhes.cn/"
        # self.verificationErrors = []
        #进入首页并登陆
        self.driver.get(self.url)
        time.sleep(1)
        self.driver.find_element_by_id("userName").send_keys("hwy")
        self.driver.find_element_by_id("password").send_keys("123456")
        self.driver.find_element_by_id("loginBtn").click()
        # 浏览器窗口最大化
        self.driver.maximize_window()

    # @classmethod
    def tearDown(self): #必须使用@classmethod装饰器, 所有case运行完之后只运行一次   tearDown():每个测试case运行完之后执行
        #关闭浏览器
        self.driver.quit()




    def isElementExist(self, element):
        '''
        捕获异常的方法判断元素是否存在
        :param element: 元素路径表达式，如Xpath "str"
        :return:
        '''
        flag = True
        driver = self.driver
        try:
            driver.find_element_by_xpath(element)
            return flag
        except:
            flag = False
            return flag


    def getliebiao(self):
        '''
        获取网页中的表格数据
        车辆管理列表为例（车辆不重复）
        :return:
        '''
        liebiao = {} #定义数组用于存储列表记录属性及属性值
        # 获取页数
        time.sleep(3)
        content_1 = self.driver.find_element_by_xpath("//div[@id='main_content']//ul/li[1]").text
        sum = int(content_1[content_1.find(" "):content_1.find("条")-1])  #“共 27 条记录 第 2 / 3 页” 获取记录总数如 27 ，截取字符串，find函数返回索引
        page = int(content_1[content_1.find("/")+1 :content_1.find("页")-1])  # 获取总页数
        # print(page)
        # 获取单页显示条数
        content_2= self.driver.find_element_by_xpath('//div[@id="main_content"]//ul/li[7]//div/div/div/div').get_attribute('title')   #获取单页显示条数：10 条/页
        page_sum = int(content_2[0:content_2.find('条')-1])
        # print(page_sum)
        # 循环页数获取页数据，一页一页读取 # //tbody/tr[n]
        for i in range(1, page+1):  # 共3页，1，2，3，按照每页读取并存储车辆
            n = 2 + i #实际页面
            u_1 = "//div[@id='main_content']//ul/li[{0}]".format(n)
            # print(u_1)
            self.driver.find_element_by_xpath(u_1).click()  #点击页数获取实际页面
            time.sleep(5)
            #获取每页的列表数据
            if i == page: #判断是否为最后一页,并计算最后一页的记录数
                page_last = sum - page_sum*(page-1)
                num = range(1, page_last+1)
            else:
                num = range(1, page_sum+1) #非最后一页按照选择的单页显示条数计算
            for j in num:  #每页10条
                u_2 = "//tbody/tr[{0}]".format(j)
                # print(u_2)
                carnum = self.driver.find_element_by_xpath(u_2 + "/td[2]").text #取每条的车牌号
                carcoulour = self.driver.find_element_by_xpath(u_2 + "/td[4]").text #车辆颜色
                cartel = self.driver.find_element_by_xpath(u_2 + "/td[5]").text  # 车辆电话
                liebiao[carnum] = {"车辆颜色":carcoulour, "车辆电话":cartel}
            # print(liebiao)
            time.sleep(5)
        print(liebiao)

    def get_date(self):
        wb = openpyxl.load_workbook("gonggao_add_data.xlsx") #获得表格文件 workbook
        sheet = wb["data1"]   #获得表单 sheet
        # 获取第一条数据，存储在字典中
        value = {}
        # print(sheet.max_row)    #获取最大行，注意行列数是从1开始计算；注意表格中是否存在空格
        for i in range(1, sheet.max_row):  # 1,2,3
            i=i+1
            # print(i)
            val_id = sheet.cell(row=i, column=1).value          #获取单元格数据；id值
            # print(val_id)
            val_title = sheet.cell(row=i, column=2).value          #title
            # print(val_title)
            val_date = sheet.cell(row=i, column=3).value           #date
            val_doc = sheet.cell(row=i, column=4).value             #doc
            value[val_id]= {"title": val_title, "data": val_date, "doc": val_doc}
            # print(value)

        return value







if __name__ == '__main__':

#      执行用例

# 执行所有
    # unittest.main()

# 一条一条添加执行
#     suit = unittest.TestSuite()
#     suit.addTest(zt_selenium_test.TestZt("test_system_gonggao_add"))
#     suit.addTest(zt_shouye_test.TestShouYe("test_go_message"))
#     runner = unittest.TextTestRunner(verbosity=2)
#     runner.run(suit)


# 调用测试用例模块，按模块执行
#     from Practices.hr_selenium0318.ZT_TestExample.zt_test.test_shouye import TestShouYe #导入用例类，注意从根目录导入
#     loader= unittest.TestLoader() #加载器实例化，否则报错
#     file_name = '../ceshibaogao.html'
#     fp = open(file_name, 'wb')
#     suit = unittest.TestSuite()  #存储器实例化
#     suit.addTest(loader.loadTestsFromTestCase(TestShouYe)) #加载并添加测试用例到存储器
#     print(suit)
#     HTMLTestRunner(stream=fp, description='测试报告描述', title='测试报告').run(suit) #执行用例
    print(TestZt().get_date())


