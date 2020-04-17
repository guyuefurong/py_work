"""
@Author  : Laura
@File    : somepoint.py
@Time    : 2020/3/9 11:25
"""

# 1. 显示当前目录
# import os
# print(os.getcwd())





# 2. 列表反向迭代排序 reverse
# list=[1,2,3,4]
# list.reverse()
# print(len(list))





# 3. 列表转换成字典
# list_ziduan=["num1","mum2","mum3","mum4"]
# zip_1=zip(list_ziduan,list) #zip()函数用于可迭代的对象作为参数，将对象中对应的元素打包成 ——》元组，返回元组组成的列表；
# 如果元素位数不一致，则返回列表长度与最短的对象相同，利用 * 号操作符，可以将元组解压为列表。
# list_ziduan=["num1","mum2","mum3"]
# zip_1=zip(list_ziduan,list)
# print(dict(zip_1)) #dict([列表1-Key，列表2-Value]) 转换为字典

# 字符串转化为列表：直接转换
# s="123456789"
# print(list(s))





# 4. 必须参数、关键字参数、默认参数、不定长参数

    # 不定长参数
    # 加" * "参数， 如*args, 加了星号 * 的参数会以元组(tuple)的形式导入，存放所有未命名的变量参数。
def get_args(*args):
    for arg in args:
        print(arg,"\n")

# 加" ** "参数， 如**kwargs, 参数带两个星号 **，传入不限个关键字参数
def get_kwargs(**kwargs):
    for key, value in kwargs.items():       #	radiansdict.items() 以列表返回可遍历的(键, 值) 元组数组
        print("{0}={1}".format(key, value),"\n")

# 如果单独出现星号 * 后的参数必须用关键字传入
def get_word(b,/,e,*,c,d,):
    res=b+c+d+e
    print(res)

# if __name__ == '__main__':
    # get_args(1, 2, 3, 4, 5, "哈哈", "lala")
    # get_kwargs(mo=1, lala=2)
    # get_word(1, 0, c=2, d=3)





# 5. 参数传递过程
#
#   不可变类型：变量赋值 a=5 后再赋值 a=10，这里实际是新生成一个 int 值对象 10，再让 a 指向它，而 5 被丢弃，不是改变a的值，相当于新生成了a。
#   可变类型：变量赋值 la=[1,2,3,4] 后再赋值 la[2]=5 则是将 list la 的第三个元素值更改，本身la没有动，只是其内部的一部分值被修改了






# 6. 数据结构 data structure
# 列表
# list=[1,8,3,3]
# list_1=[]
# list.append(1) #追加
# list.extend(list) #添加指定列表的所有元素来扩充列表
# list.insert(0, "mm") #指定位置插入，示例为0位置插入“mm”
# list.index(3) #返回列表中第一个值为 x 的元素的索引
# list.remove(1) #	删除列表中值为 x 的第一个元素。如果没有这样的元素，就会返回一个错误。
# list.pop(1) #从列表的指定位置移除元素，并将其返回。删除1 位置的元素并返回该元素,不是删除原始数据
# list.count(1) #统计元素出现次数
# list.sort() #对列表中的元素进行排序，返回none
# print(list)
# list.reverse() #倒排列表中的元素
# list_1=list.copy() #返回列表的浅复制
# 内置函数，去重：set(list)
# list_1=list.pop(1)
# print(list)
# print(set(list))
# for i in range(9,3,-1):
#     print(i)




# 7. 字典遍历：
# 在字典中遍历时，关键字和对应的值可以使用 items() 方法同时解读出来：
# knights = {'gallahad': 'the pure', 'robin': 'the brave'}
# for k, v in knights.items():
#     print(k, v)

# 在序列中遍历时，索引位置和对应值可以使用 enumerate() 函数同时得到
# for i, v in enumerate(['tic', 'tac', 'toe']):
#     print(i, v)

# 同时遍历两个或更多的序列，可以使用 zip() 组合
# questions = ['name', 'quest', 'favorite color']
# answers = ['lancelot', 'the holy grail', 'blue']
# for q, a in zip(questions, answers):
#     print('What is your {0}?  It is {1}.'.format(q, a))

# 要反向遍历一个序列，首先指定这个序列，然后调用 reversed() 函数
# for i in reversed(range(1, 10, 2)):
#     print(i)


# 要按顺序遍历一个序列，使用 sorted() 函数返回一个已排序的序列，并不修改原值，set()set() 函数创建一个无序不重复元素集（去重）
# basket = ['apple', 'orange', 'apple', 'pear', 'orange', 'banana']
# for f in sorted(set(basket)):
#     print(f)






# 8. 递归：程序调用自身的编程技巧称为递归（ recursion）。一般来说，递归需要有边界条件、递归前进段和递归返回段。当边界条件不满足时，递归前进；当边界条件满足时，递归返回。
# def count(n):
#     '''计算n的阶乘'''
#     if n==1 or n==0:
#         return 1
#     else:
#         print(n)
#         return n*count(n-1)
#
# if __name__ == '__main__':
#     k=int(input('请输入一个整数')) #k=4
#     res=count(k)
#     print(res)

##尾递归优化,尾递归是指，在函数返回的时候，调用自身本身，并且，return语句不能包含表达式。
# 这样，编译器或者解释器就可以把尾递归做优化，使递归本身无论调用多少次，都只占用一个栈帧，不会出现栈溢出的情况。
# def count(n,res=1):
#     if n==1 or n==0:
#         return res
#     print(n, res)
#     return count(n-1,n*res)
#
# print(count(4))

# def recursion(n):
#     v = n//2 # 地板除，保留整数
#     print(v) # 每次求商，输出商的值
#     if v==0:
#         ''' 当商为0时，停止，返回Done'''
#         return 'Done'
#     v = recursion(v)



# 9.类/函数/对象
# 解决接口键关联性问题，如下个请求需要用到上个请求的数据，cookies等
# method1：使用unittest 框架中初始化函数 setUp() ，每次执行用例前执行获取cookies
# method2：使用全局变量，设置全局变量初始值，并修改初始值，每次使用时都需声明全局变量（比较重复）
# method3：使用映射，模块之外新建类函数，使用时导入函数，通过调用修改/获取/删除等方法来使用


# 10.python 对表格实现读写操作（可进行参数化运用）
#   方法一：openpyxl 模块实现读写（修改）
from openpyxl import Workbook
from openpyxl import load_workbook
# from openpyxl.writer.excel import ExcelWriter

# '''
# 拷贝 sheet
# '''
# # 读取 somepoint_date.xlsx
# wb = load_workbook(u"somepoint_data.xlsx")
# # 获取当前的 sheet
# source = wb.active
# # 拷贝 sheet
# target = wb.copy_worksheet(source)
# # 给 sheet 重命名
# target.title = 'copy'
# # 保存
# wb.save(u"somepoint_data.xlsx")
#
#
# '''
# 修改值
# '''
# wb = load_workbook(u"somepoint_data.xlsx")
# # 获取所有 sheet 名称
# sheetnames =wb.sheetnames     #获取工作表所有表单名称，get_sheet_names函数python3已弃用
# # 获取第一个 sheet
# sheet = wb[sheetnames[0]]     #sheetnames[0]表示按照索引取第一张表单
# # A 列求和，并赋值到 D1
# # sheet["D1"] = "=SUM(A:A)"
# # 打印 A2 的值
# print(sheet['A2'].value)
# # 打印位于第6行，第3列的值
# print(sheet.cell(row=2,column=2).value)
# # 赋值给 A1
# sheet['A1'] = '47'
# # 保存
# wb.save("somepoint_data.xlsx")
#
#
# '''
# 创建 excel
# '''
# # 创建一个 excel
# wb = Workbook() #创建的表格不能再此处传参命名，只能再保存时命名
# # 获取当前的 sheet
# sheet = wb.active
# # 赋值 A1 为 4
# sheet['A1'] = 4
# # 保存
# wb.save("create.xlsx")


'''
插入空行
'''
# 读取 output.xlsx
wb = load_workbook('output.xlsx')
sheet = wb.copy_worksheet(wb.active)
sheet.title = 'copy'
# 在第二行(idx)上方插入两个(amount)空行
sheet.insert_rows(idx=2, amount=2)
# 保存
wb.save('insert.xlsx')








    # print(1)
    # sheet1=wb.get_sheet_by_name(name)   #获取表单名称



#   方法二：xlrd (xlsx表格read） 和 xlwt （xlsx表格write）模块实现
#   方法三：xlsxwrite 模块实现

# rw_data("somepoint_data.xlsx","data")






