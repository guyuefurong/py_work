"""
@Author  : Laura
@File    : Tools.py
@Time    : 2020/4/16 17:50
"""
"""
封装测试常用工具
"""
import time
import openpyxl


def s_time(num):
    '''
    等待时间
    :param num: 设置的等待时长，单位秒(/s)
    :return: None
    '''
    time.sleep(num)

def get_data(xlsx):
    '''
    读取驾驶员信息表格数据
    :param xlsx: 表名
    :return: 返回 list类型数据
    '''
    wb=openpyxl.load_workbook(xlsx)   #获取表格
    sheet= wb["data"]       #定位表单
    list = []
    # print(sheet.max_row)    #获取最大行，注意行列数是从1开始计算；注意表格中是否存在空格
    for i in range(1, sheet.max_row):  # 1,2,3
        i = i + 1
        # print(i)
        val_id = sheet.cell(row=i, column=1).value  # 获取单元格数据；id值
        val_name = sheet.cell(row=i, column=2).value
        if val_name not in list:
            list.append(val_name)
    # print(list)
    # print(type(list))
    return list